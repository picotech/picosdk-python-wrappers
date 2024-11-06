#
# Copyright (C) 2018 Pico Technology Ltd. See LICENSE file for terms.
#

import numpy as np
import openpyxl
from math import floor, log2, log10
import shutil
import os
import toml

def dataImporter(name):

    workbook = openpyxl.load_workbook(name, data_only = True)
    
    sheet = workbook['filterParameters']
    
    noOfChannels = sheet.cell(row = 2, column = 1).value
    bits = sheet.cell(row = 2, column = 2).value
    samplingRate = sheet.cell(row = 2, column = 3).value
    sampleLength = sheet.cell(row = 2, column = 4).value
    
    return noOfChannels, bits, samplingRate, sampleLength
    
    
def ps6000aTimebase(samplingRate):
    
    sampleInterval = (1/samplingRate)/1000000 #s
    
    breakPoint = 6.4/1000000000
    
    if sampleInterval >= breakPoint:
        timebase = floor((sampleInterval * 156250000)+5)
    else:
        timebase = floor(log2(sampleInterval * 5000000000))
    
    return timebase
    
def ps5000aTimebase(samplingRate):
    
    sampleInterval = (1/samplingRate)/1000000 #s
    
    breakPoint = 8/1000000000
    
    if sampleInterval >= breakPoint:
        timebase = floor((sampleInterval * 125000000)+2)
    else:
        timebase = floor(log2(sampleInterval * 1000000000))
    
    return timebase
    
def ps3000aTimebase(samplingRate):
    
    sampleInterval = (1/samplingRate)/1000000 #s
    
    breakPoint = 8/1000000000
    
    if sampleInterval >= breakPoint:
        timebase = floor((sampleInterval * 125000000)+2)
    else:
        timebase = floor(log2(sampleInterval * 1000000000))
    
    return timebase
  
def ps4000aTimebase(samplingRate):
    
    timebase = floor((80/samplingRate)-1)
    
    return timebase
    
def ps2000aTimebase(samplingRate):
    
    sampleInterval = (1/samplingRate)/1000000 #s
    
    breakPoint = 4/1000000000
    
    if sampleInterval>= breakPoint:
        timebase = floor((sampleInterval*125000000)+2)
    else:
        timebase = floor(log2(sampleInterval * 1000000000))
        
    return timebase
    
def ps2000Timebase(sampleRate):
    #assumes sample rate is in Hz
    #assumes sample interval in s
    
    sampleInterval = (1/sampleRate)
    
    timebase = floor(log10(sampleInterval*1000000000))
    
    return timebase
    
    
def BitEnumSelector(bits):

    if bits <= 8:
        enum = 0
    elif bits <= 10:
        enum = 10
    else:
        enum = 1
        
    return enum
    
def saveConfigFile(channels, bits, sampleRate,captureLength, segments):
    
    # configValues = [channels, bits, sampleRate, captureLength, segments]
    data = {
        "Active Channels" : int(channels),
        "Scope Bit Resolution" : int(bits),
        "Sampling Rate (MHz)" : float(sampleRate),
        "Capture Length (Samples)" : int(captureLength),
        "Number of Capture Segments for Rapid Block" : int(segments),
        }
    # # Save the list to a text file
    with open('configValues.toml', 'w') as file:
        toml.dump(data,file)  
    
        
    return
    
def loadConfigValues():
    
    with open('configValues.toml', 'r') as file:
        restored_configValues = toml.load(file)
            
    channels = int(restored_configValues["Active Channels"])
    bits = int(restored_configValues["Scope Bit Resolution"])
    sampleRate = float(restored_configValues["Sampling Rate (MHz)"])
    captureLength = int(float(restored_configValues["Capture Length (Samples)"]))
    segments = int(restored_configValues["Number of Capture Segments for Rapid Block"])
    
    return channels, bits, sampleRate, captureLength, segments
    
def copyFile(source_directory, filename):
    
    destination_directory = 'D:/'

    # Define the destination file path
    source_file = os.path.join(source_directory, filename)
    destination_file = os.path.join(destination_directory, filename)

    # Copy the file to the new location
    shutil.copy(source_file, destination_file)
    
    return 